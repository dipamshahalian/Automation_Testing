import openpyxl

# File->Workbook->Sheets->Rows->Cells
file = "C:/Users/Alian-172/Downloads/dataset2.xlsx"

workbook = openpyxl.load_workbook(file)

sheet = workbook["Sheet1"]

rows = sheet.max_row # Count no. of rows

cols = sheet.max_column #no. of cols.

# Reading all rows and cols
for r in range(1, rows+1):
    for c in range(1, cols+1):
        print(sheet.cell(r, c).value, end = "           ")
    print()