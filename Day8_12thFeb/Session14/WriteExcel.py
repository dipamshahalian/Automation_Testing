import openpyxl

file = "C:/Users/Alian-172/Downloads/dataset2.xlsx"

workbook = openpyxl.load_workbook(file)

sheet = workbook["Sheet1"]
# sheet = workbook["Sheet_Name"]
# sheet = workbook.active

# for r in range(1, 11):
#     for c in range(1, 9):
#         sheet.cell(r, c).value = "Dipam"

# Above is used to print the same data in every cell

#  Now for multiple data
sheet.cell(1, 1).value = 123
sheet.cell(1, 2).value = "Smith"

sheet.cell(2, 1).value = 124
sheet.cell(2, 2).value = "John"

sheet.cell(3, 1).value = 125
sheet.cell(3, 2).value = "Dipam"


workbook.save(file)