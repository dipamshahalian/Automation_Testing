from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Load Excel file
file_path = "C:/Users/Alian-172/Downloads/Logindata1.xlsx"
wb = load_workbook(file_path)
ws = wb.active  # Active sheet

# Read login credentials from Excel
email = ws.cell(row=2, column=1).value  # First row, first column (Email)
password = ws.cell(row=2, column=2).value  # First row, second column (Password)

# Set up Selenium WebDriver
service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service)

try:
    # Open Login Page
    driver.get("https://alianhub-beta-2lu4z.ondigitalocean.app/#/login?redirect_url=/67a1d65540659a99bfbe83db/settings/my-profile")
    driver.maximize_window()

    # Enter Email
    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, "email"))
    ).send_keys("aliantest236+4020@gmail.com")

    # Enter Password
    driver.find_element(By.ID, "password").send_keys("Abc@223133")

    # Click Login Button
    driver.find_element(By.XPATH, "//button[@type='submit']").click()

    # Wait for Successful Login Check
    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "//span[contains(text(),'My Profile')]"))  # Update with a valid element that confirms login
    )

    # Update Excel with "Passed"
    ws.cell(row=2, column=3, value="Passed")
    print("Login successful. Status updated in Excel.")

except Exception as e:
    # If login fails, update Excel with "Failed"
    ws.cell(row=2, column=3, value="Failed")
    print(f"Login failed: {e}")

# Save and close Excel
wb.save("C:/Users/Alian-172/Downloads/Logindata1.xlsx")
wb.close()

# Close browser
driver.quit()