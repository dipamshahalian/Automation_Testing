from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
import time
from openpyxl import load_workbook

service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service)
driver.get("https://alianhub-beta-2lu4z.ondigitalocean.app/#/login?redirect_url=/business")
driver.maximize_window()

driver.find_element(By.XPATH, "//input[@id='email']").send_keys("aliantest236+4020@gmail.com")
driver.find_element(By.XPATH, "//input[@id='password']").send_keys("Abc@223133")
driver.find_element(By.XPATH, "//button[@type='submit']").click()
time.sleep(10)

No_button = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.XPATH, "//button[text()='No']"))
)
No_button.click()

driver.find_element(By.XPATH, "//a[normalize-space()='Projects']").click()
time.sleep(2)
driver.find_element(By.XPATH, "//button[@id='createproject_driver']").click()
time.sleep(1)
driver.find_element(By.XPATH, "//div[@id='createprojectusingtemplate_driver']//button[@type='button']").click()
time.sleep(1)
driver.find_element(By.XPATH, "//body/div[@id='app']/div[@id='my-sidebar']/div[@class='position-fi justify-content-end sidebar d-flex']/div[@id='createprojectfirst_driver']/div[@class='black sidebar-body bg-white style-scroll']/div[@class='bg-light-gray mobile-common-background h-100 addUseTemplate']/div[@class='header-sidebar default-background-header d-flex bg-white border-radius-8-px']/div[@id='create-project-loading']/div[@class='statusTaskWrapper statusTaskWrapperMain usetemplatesWrapper d-flex useTemaplteHeight']/div[@class='template-right-side position-re']/div/div/div[@class='d-flex mobile-usetempllate-list-display w-100 flex-wrap']/div[1]").click()
time.sleep(1)
driver.find_element(By.XPATH, "//button[@type='button']").click()

wb = load_workbook("C:/Users/Alian-172/Downloads/Logindata.xlsx")
ws = wb.active

project_name = ws.cell(row=1, column=1).value  # First row, first column
key = ws.cell(row=1, column=2).value  # First row, second column

driver.find_element(By.XPATH, "//div[@id='createprojectname_driver']//input[@id='inputId']").send_keys(project_name)
driver.find_element(By.XPATH, "//div[@id='createprojectkey_driver']//input[@id='inputId']").send_keys(key)
driver.find_element(By.XPATH, "//div[@id='createprojectcategory_driver']//input[@id='inputId']").click()
time.sleep(2)
driver.find_element(By.XPATH, "//div[@id='item2']").click()
time.sleep(1)
driver.find_element(By.XPATH, "//button[@class='submit-btn templateall__submit-btn cursor-pointer conditional-submit btn']").click()
print("New Project Created")

time.sleep(150)

driver.quit()